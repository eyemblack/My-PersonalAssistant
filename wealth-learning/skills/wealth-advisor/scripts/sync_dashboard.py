import openpyxl
import json
import os
import re
import urllib.request
import ssl

GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1IR2W8lKU6OT7YyJbCNAYZ62CeD41l77Vo0uV2p56CRA/export?format=xlsx"
EXCEL_PATH = '/Users/opentechbox/Desktop/PA/wealth-learning/Income_Expenses.xlsx'
MONTHLY_PLAN_HTML = '/Users/opentechbox/Desktop/PA/wealth-learning/monthly_plan.html'
LAYOFF_SIMULATOR_HTML = '/Users/opentechbox/Desktop/PA/wealth-learning/layoff_simulator.html'

def sync_all():
    # 0. Download latest sheet from Google Cloud
    try:
        print("Downloading latest Google Sheet from cloud...")
        # Bypass SSL Verification for macOS Python URL opens
        ssl_ctx = ssl._create_unverified_context()
        req = urllib.request.Request(GOOGLE_SHEET_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, context=ssl_ctx) as response:
            with open(EXCEL_PATH, 'wb') as f:
                f.write(response.read())
        print("Download completed successfully!")
    except Exception as e:
        print(f"Warning: Could not download latest sheet from cloud ({e}). Using local copy.")

    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return
        
    print("Loading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # 1. Parse Data
    m_data = parse_monthly_plan(wb)
    o_data = parse_overall(wb)
    
    # 2. Update monthly_plan.html
    if os.path.exists(MONTHLY_PLAN_HTML):
        update_monthly_plan(m_data, o_data)
    else:
        print("Warning: monthly_plan.html not found.")
        
    # 3. Update layoff_simulator.html
    if os.path.exists(LAYOFF_SIMULATOR_HTML):
        update_layoff_simulator(m_data, o_data)
    else:
        print("Warning: layoff_simulator.html not found.")
        
    print("Dashboard synchronization completed successfully!")

def parse_monthly_plan(wb):
    sheet = wb["Monthly Plan"]
    
    # Income (Net)
    net_income = sheet["F2"].value or 47793.22
    
    # Read fixed expenses (Row 1 to 14, column A & B)
    fixed_expenses = []
    fixed_sum = 0
    for r in range(1, 15):
        name = sheet.cell(row=r, column=1).value
        val = sheet.cell(row=r, column=2).value
        if name and name != "SUM" and val is not None:
            fixed_expenses.append({"name": name, "val": val})
            fixed_sum += val
                
    # Read remaining budget allocations (Row 22 to 30)
    scenarios = {
        "current": [],
        "july": [],
        "nov": [],
        "y2027": []
    }
    
    col_mapping = {
        2: "current",
        3: "july",
        4: "nov",
        5: "y2027"
    }
    
    method_mapping = {
        "ผ่อน IP 17Pro": "จ่ายค่าเครื่องผ่อนชำระ",
        "เที่ยว, กิน kept": "บัญชี Kept (ท่องเที่ยว/กินเล่น)",
        "ภาษี kept": "บัญชี Kept (สะสมยื่นภาษีปี)",
        "เงินฉุกเฉิน kept": "บัญชี Kept (สะสมเงินสำรอง)",
        "กยศ dime": "แอป Dime (ออมรอตัดยอดรายปี)",
        "รถยนต์ kept": "บัญชี Kept (บำรุงรักษารถ)",
        "รอเทรดหุ้น dime": "แอป Dime (รอจังหวะเทรดหุ้น)",
        "Gemini kept": "บัญชี Kept (รายเดือน AI)",
        "ประกันรวม kept": "บัญชี Kept (สะสมค่าประกัน)"
    }
    
    for r in range(22, 31):
        name = sheet.cell(row=r, column=1).value
        if not name or name == "รวม":
            continue
        for c in [2, 3, 4, 5]:
            scen_key = col_mapping[c]
            val = sheet.cell(row=r, column=c).value or 0.0
            
            # map readable descriptions
            method = method_mapping.get(name, "บัญชี Kept")
            if name == "ผ่อน IP 17Pro" and val == 0:
                method = "หมดหนี้ผ่อนชำระโทรศัพท์"
            elif name == "รอเทรดหุ้น dime" and val > 0:
                method = "แอป Dime ( DCA หุ้นต่างประเทศ )"
            
            scenarios[scen_key].append({
                "name": name,
                "method": method,
                "val": val
            })
            
    # Read OT splits (Row 37-38)
    ot_splits = {}
    for c in range(2, 6):
        name = sheet.cell(row=37, column=c).value
        val = sheet.cell(row=38, column=c).value
        if name:
            ot_splits[name.strip()] = val
            
    # Read Bonus plan (Row 44 to 56)
    bonus_items = []
    for r in range(45, 54):
        name = sheet.cell(row=r, column=1).value
        val = sheet.cell(row=r, column=2).value
        note = sheet.cell(row=r, column=3).value or ""
        if name and name != "รวม" and val is not None:
            bonus_items.append({"name": name, "val": val, "note": note})
            
    return {
        "net_income": net_income,
        "fixed_expenses": fixed_expenses,
        "fixed_sum": fixed_sum,
        "scenarios": scenarios,
        "ot_splits": ot_splits,
        "bonus_items": bonus_items
    }

def parse_overall(wb):
    sheet = wb["Overall"]
    
    contingency = {
        "kbank": sheet["D13"].value or 1000.0,
        "kept": sheet["D14"].value or 108380.12,
        "blb": sheet["D15"].value or 0.0,
        "dime": sheet["D16"].value or 10000.0
    }
    contingency["sum"] = sum(contingency.values())
    
    hold_for_pay = {
        "tax": sheet["D24"].value or 13008.06,
        "studentloan": sheet["D25"].value or 8000.0,
        "travel": sheet["D26"].value or 21530.72,
        "car": sheet["D27"].value or 27597.32,
        "bigthing": sheet["D28"].value or 3494.18,
        "learn": sheet["D29"].value or 284.0,
        "insure": sheet["D30"].value or 11964.0,
        "subscribe": sheet["D31"].value or 7500.0
    }
    hold_for_pay["sum"] = sum(hold_for_pay.values())
    
    total_kept = contingency["kept"] + hold_for_pay["sum"]
    total_networth = contingency["sum"] + hold_for_pay["sum"]
    
    return {
        "contingency": contingency,
        "hold_for_pay": hold_for_pay,
        "total_kept": total_kept,
        "total_networth": total_networth
    }

def update_monthly_plan(m_data, o_data):
    with open(MONTHLY_PLAN_HTML, 'r', encoding='utf-8') as f:
        content = f.read()
        
    net_str = f"{m_data['net_income']:,.2f}"
    content = re.sub(
        r'<span style="color: var\(--green\); font-size: 1.15rem;">.*?</span>',
        f'<span style="color: var(--green); font-size: 1.15rem;">{net_str} บาท</span>',
        content
    )
    
    scenarios_js = "const scenarios = " + json.dumps({
        "current": {
            "title": "แผนปัจจุบัน (Current Case)",
            "note": "เน้นการเริ่มผ่อนชำระหนี้พิเศษอย่างครอบคลุม และยังเก็บออมเงินสำรองฉุกเฉินเบื้องต้น 5,000 บาท/เดือน มีค่าผ่อนโทรศัพท์ IP 17Pro 3,880 บาท/เดือน ซึ่งจะหมดลงในอนาคตอันใกล้",
            "items": m_data["scenarios"]["current"]
        },
        "july": {
            "title": "แผนเดือนกรกฎาคม (July)",
            "note": "เงินกองฉุกเฉินถูกยกระดับขึ้นเป็น 10,412.22 บาท/เดือน (เพิ่มเกือบเท่าตัว) โดยลดการสะสมเงินท่องเที่ยวและงดส่งเงินเข้ากอง กยศ ชั่วคราว (เนื่องจากมีการสำรองครบยอดจ่ายปีนี้แล้ว หรือนำไปเติมในกองฉุกเฉินให้ได้ไวที่สุดก่อน)",
            "items": m_data["scenarios"]["july"]
        },
        "nov": {
            "title": "แผนเดือนพฤศจิกายน (Nov - หลังหมดงวดผ่อนโทรศัพท์)",
            "note": "<strong>จุดเปลี่ยนสำคัญ:</strong> โทรศัพท์ผ่อนหมดแล้ว ทำให้ยอด 3,880 บาทกลายเป็นเงินอิสระ! สังเกตว่าเงินออมฉุกเฉินพุ่งขึ้นสู่จุดสูงสุดที่ 17,292.22 บาท/เดือน และคงงบเที่ยวไว้ที่ 2,000 บาทเพื่อไม่ให้การท่องเที่ยวอึดอัดจนเกินไป",
            "items": m_data["scenarios"]["nov"]
        },
        "y2027": {
            "title": "แผนระยะยาว ปี 2027 (ระบบเสถียร)",
            "note": "เมื่อหนี้สั้นหมดลงและเงินสำรองฉุกเฉินมีความมั่นคงขึ้น คุณจะปรับพอร์ตเข้าสู่ Level 3 อย่างแท้จริง โดยแผนล่าสุดคุณเลือกที่จะ<strong>เก็บเงินสำรองฉุกเฉินเพิ่มเป็น 15,192.22 บาท/เดือน</strong> เพื่อสร้างฐานรากที่ปลอดภัยอย่างยาวนานตามสถานการณ์ปัจจุบัน",
            "items": m_data["scenarios"]["y2027"]
        }
    }, indent=4, ensure_ascii=False) + ";"
    
    content = re.sub(
        r'const scenarios = \{.*?\};',
        scenarios_js,
        content,
        flags=re.DOTALL
    )
    
    with open(MONTHLY_PLAN_HTML, 'w', encoding='utf-8') as f:
        f.write(content)
    print("Updated monthly_plan.html successfully.")

def update_layoff_simulator(m_data, o_data):
    with open(LAYOFF_SIMULATOR_HTML, 'r', encoding='utf-8') as f:
        content = f.read()
        
    content = re.sub(r'id="param-salary" value="\d+"', f'id="param-salary" value="{int(m_data["net_income"])}"', content)
    content = re.sub(r'id="param-cash" value="\d+"', f'id="param-cash" value="{int(o_data["contingency"]["sum"])}"', content)
    
    with open(LAYOFF_SIMULATOR_HTML, 'w', encoding='utf-8') as f:
        f.write(content)
    print("Updated layoff_simulator.html successfully.")

if __name__ == '__main__':
    sync_all()
