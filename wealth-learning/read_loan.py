import openpyxl

wb = openpyxl.load_workbook('/Users/opentechbox/Desktop/PA/wealth-learning/Income_Expenses.xlsx', data_only=True)
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"--- Sheet: {sheetname} ---")
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None and any(word in str(val) for word in ["กรอ", "กยศ", "Student", "student"]):
                print(f"Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)}{r}): {val}")
                # Print adjacent cells to see the context/amount
                context = []
                for adj_c in range(max(1, c - 2), min(sheet.max_column + 1, c + 5)):
                    adj_val = sheet.cell(row=r, column=adj_c).value
                    context.append(f"{openpyxl.utils.get_column_letter(adj_c)}{r}: {adj_val}")
                print("  Context:", ", ".join(context))
